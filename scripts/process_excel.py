#!/usr/bin/env python3
"""
국토부 실거래가 엑셀 → properties.json + prices.json 변환 스크립트

사용법:
  python3 scripts/process_excel.py [--row-data ./rowData] [--output ./entities/property/model] [--no-geocoding]

필요 패키지:
  pip install openpyxl
"""

import argparse
import json
import os
import glob
import re
import statistics
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
import warnings

# openpyxl 스타일 경고 무시
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def nfc(s: str) -> str:
    """macOS NFD 파일명을 NFC로 정규화"""
    return unicodedata.normalize('NFC', s)

HEADER_ROW = 13   # 모든 파일 공통 헤더 행 번호
DATA_START_ROW = 14


# ──────────────────────────────────────────
# 유틸리티 함수
# ──────────────────────────────────────────

def sqm_to_pyeong(sqm: float) -> float:
    """㎡ → 평 (소수점 1자리)"""
    return round(float(sqm) * 0.3025, 1)


def sqm_to_pyeong_group(sqm: float) -> int:
    """
    ㎡ → 평형 그룹 (정수 반올림)
    예) 30.2평, 30.4평, 30.8평 → 모두 30평형으로 통합
    실제 아파트 공급면적 기준으로 ±2평 이내는 같은 평형으로 취급
    """
    pyeong = float(sqm) * 0.3025
    return round(pyeong)


def parse_price(price_val) -> Optional[int]:
    """'263,000' 또는 263000.0 → 2630000000 (원). 유효하지 않으면 None."""
    if price_val is None:
        return None
    s = str(price_val).replace(',', '').strip()
    if s in ['-', '', 'None', '0']:
        return None
    try:
        return int(float(s)) * 10000
    except ValueError:
        return None


def make_property_id(name: str, sigungu: str, ptype: str) -> str:
    """
    property ID 생성: apt-{단지명}-{구동} 또는 opt-{단지명}-{구동}
    예) apt-시범삼성-분당구서현동
    """
    prefix = 'apt' if ptype == 'apartment' else 'opt'
    # 시군구에서 "경기도 성남시 " 제거 후 공백 제거
    district = re.sub(r'경기도\s*성남시\s*', '', sigungu).replace(' ', '')
    clean_name = re.sub(r'[^가-힣a-zA-Z0-9]', '', name)
    return f'{prefix}-{clean_name}-{district}'


def make_unit_id(property_id: str, area_sqm: float) -> str:
    """
    유닛 ID: {property_id}-{평형그룹}평형
    층수를 제외하고 면적만으로 그룹핑 → 같은 단지의 30.2/30.4/30.8평 → 모두 30평형으로 통합
    """
    pyeong_group = sqm_to_pyeong_group(area_sqm)
    return f'{property_id}-{pyeong_group}평형'


def to_str(val) -> str:
    """None-safe 문자열 변환"""
    if val is None:
        return ''
    return str(val).strip()


def to_int(val, default: int = 0) -> int:
    """None-safe 정수 변환"""
    if val is None:
        return default
    try:
        return int(float(str(val).replace(',', '').strip()))
    except (ValueError, TypeError):
        return default


def to_float(val, default: float = 0.0) -> float:
    """None-safe 실수 변환"""
    if val is None:
        return default
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return default


# ──────────────────────────────────────────
# 엑셀 파일 파싱
# ──────────────────────────────────────────

def read_excel(filepath: str) -> List[dict]:
    """엑셀 파일을 읽어서 딕셔너리 리스트로 반환 (헤더 기준)"""
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb.active

    # 헤더 읽기 (13행)
    max_col = ws.max_column
    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, max_col + 1)]

    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        # 모두 None이면 빈 행 → 스킵
        if not any(v for v in row_vals if v is not None):
            continue
        row_dict = {}
        for h, v in zip(headers, row_vals):
            if h is not None:
                row_dict[h] = v
        rows.append(row_dict)

    wb.close()
    return rows


def detect_file_type(filename: str) -> str:
    """파일명으로 타입 자동 감지 (macOS NFD 유니코드 정규화 처리)"""
    fn = nfc(os.path.basename(filename))
    if '아파트' in fn and '매매' in fn:
        return 'apt_sale'
    elif '아파트' in fn and ('전월세' in fn or '전세' in fn):
        return 'apt_jeonse'
    elif '오피스텔' in fn and '매매' in fn:
        return 'opt_sale'
    elif '오피스텔' in fn and ('전월세' in fn or '전세' in fn):
        return 'opt_jeonse'
    raise ValueError(f'알 수 없는 파일 유형: {filename}')


# ──────────────────────────────────────────
# 집계 로직
# ──────────────────────────────────────────

def aggregate(
    sale_rows: List[dict],
    jeonse_rows: List[dict],
    ptype: str,
    existing_coord_map: Dict[str, Tuple[float, float]],
) -> Tuple[List[dict], List[dict], dict]:
    """
    매매/전월세 데이터를 집계하여 properties와 prices 리스트 반환.

    Returns:
        properties: list[dict]
        prices: list[dict]
        stats: 통계 딕셔너리
    """
    buildings: Dict[str, dict] = {}   # property_id → property
    units: Dict[str, dict] = {}       # unit_id → unit (가격 집계용)

    skipped_cancel = 0
    skipped_no_name = 0
    skipped_no_price = 0
    total_sale = 0

    # ── 1. 매매 데이터로 시세 집계 ──
    for row in sale_rows:
        total_sale += 1

        # 해제된 거래 건너뜀
        cancel = to_str(row.get('해제사유발생일', '-'))
        if cancel not in ['-', '', 'None']:
            skipped_cancel += 1
            continue

        name = to_str(row.get('단지명', ''))
        if not name:
            skipped_no_name += 1
            continue

        sigungu = to_str(row.get('시군구', ''))
        area_sqm = to_float(row.get('전용면적(㎡)', 0))
        floor = to_int(row.get('층', 1), default=1)
        build_year = to_int(row.get('건축년도', 2000), default=2000)
        road = to_str(row.get('도로명', ''))
        year_month = to_str(row.get('계약년월', ''))

        price = parse_price(row.get('거래금액(만원)'))
        if not price:
            skipped_no_price += 1
            continue

        prop_id = make_property_id(name, sigungu, ptype)
        area_pyeong = sqm_to_pyeong(area_sqm)
        pyeong_group = sqm_to_pyeong_group(area_sqm)
        unit_id = make_unit_id(prop_id, area_sqm)

        # property 등록
        if prop_id not in buildings:
            address = sigungu
            addr_parts = sigungu.rsplit(' ', 1)
            road_address = f'{addr_parts[0]} {road}'.strip() if road else address

            coord_key = f'{name}|{sigungu}'
            lat, lng = existing_coord_map.get(coord_key, (0.0, 0.0))
            if lat == 0.0 and lng == 0.0:
                lat, lng = existing_coord_map.get(prop_id, (0.0, 0.0))

            buildings[prop_id] = {
                'id': prop_id,
                'name': name,
                'address': address,
                'roadAddress': road_address,
                'type': ptype,
                'lat': lat,
                'lng': lng,
                'buildYear': build_year,
                'units': [],
            }

        # unit 등록 (평형 그룹 기준 — 30.2/30.4/30.8평 모두 30평형으로 합산)
        if unit_id not in units:
            units[unit_id] = {
                'id': unit_id,
                'propertyId': prop_id,
                'area': area_sqm,          # 대표 실거래 면적 (첫 번째 거래값)
                'areaPyeong': float(pyeong_group),  # 그룹 평형 (정수)
                'officialPrice': 0,
                '_marketPrices': [],
                '_jeonsePrices': [],
                '_monthlyRents': [],
                'lastTransactionDate': year_month,
            }
            if unit_id not in buildings[prop_id]['units']:
                buildings[prop_id]['units'].append(unit_id)

        units[unit_id]['_marketPrices'].append(price)
        # 최신 거래일 갱신
        if to_str(year_month) > to_str(units[unit_id]['lastTransactionDate']):
            units[unit_id]['lastTransactionDate'] = year_month

    # ── 2. 전월세 데이터로 전세가/월세 집계 ──
    for row in jeonse_rows:
        name = to_str(row.get('단지명', ''))
        if not name:
            continue

        sigungu = to_str(row.get('시군구', ''))
        area_sqm = to_float(row.get('전용면적(㎡)', 0))
        floor = to_int(row.get('층', 1), default=1)
        contract_type = to_str(row.get('전월세구분', ''))

        deposit = parse_price(row.get('보증금(만원)'))
        monthly = parse_price(row.get('월세금(만원)'))

        prop_id = make_property_id(name, sigungu, ptype)
        unit_id = make_unit_id(prop_id, area_sqm)

        if unit_id not in units:
            continue  # 매매 데이터가 없는 유닛은 스킵

        if contract_type == '전세' and deposit:
            units[unit_id]['_jeonsePrices'].append(deposit)
        elif contract_type == '월세' and monthly:
            units[unit_id]['_monthlyRents'].append(monthly)

    # ── 3. 집계 → 최종 prices 생성 ──
    prices = []
    for unit_id, unit in units.items():
        market_prices = unit.pop('_marketPrices', [])
        jeonse_prices = unit.pop('_jeonsePrices', [])
        monthly_rents = unit.pop('_monthlyRents', [])

        unit['marketPrice'] = int(statistics.median(market_prices)) if market_prices else 0
        unit['jeonsePrice'] = int(statistics.median(jeonse_prices)) if jeonse_prices else 0
        unit['monthlyRent'] = int(statistics.median(monthly_rents)) if monthly_rents else 0
        unit['transactionCount'] = len(market_prices)

        # lastTransactionDate: YYYYMM → "YYYY-MM" 형식
        ldt = to_str(unit.get('lastTransactionDate', ''))
        if len(ldt) == 6 and ldt.isdigit():
            unit['lastTransactionDate'] = f'{ldt[:4]}-{ldt[4:6]}'
        else:
            unit['lastTransactionDate'] = ldt

        prices.append(unit)

    properties = list(buildings.values())

    stats = {
        'total_sale_rows': total_sale,
        'skipped_cancel': skipped_cancel,
        'skipped_no_name': skipped_no_name,
        'skipped_no_price': skipped_no_price,
        'properties': len(properties),
        'units': len(prices),
        'no_coord': sum(1 for p in properties if p['lat'] == 0.0 and p['lng'] == 0.0),
    }

    return properties, prices, stats


# ──────────────────────────────────────────
# 기존 좌표 맵 로드 (이름+시군구 → lat, lng)
# ──────────────────────────────────────────

def load_existing_coords(props_path: Path) -> Dict[str, Tuple[float, float]]:
    """
    기존 properties.json에서 좌표 맵을 로드.
    키: '{name}|{siGunGu}' 또는 새 형식 ID
    """
    coord_map: Dict[str, Tuple[float, float]] = {}
    if not props_path.exists():
        return coord_map

    try:
        with open(props_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f'  경고: 기존 properties.json 로드 실패: {e}')
        return coord_map

    if isinstance(data, list):
        for item in data:
            lat = float(item.get('lat', 0) or 0)
            lng = float(item.get('lng', 0) or 0)
            if lat == 0.0 and lng == 0.0:
                continue

            # 기존 구조: siGunGu 필드
            name = item.get('name', '')
            sigungu = item.get('siGunGu', '') or item.get('address', '')
            if name and sigungu:
                coord_map[f'{name}|{sigungu}'] = (lat, lng)

            # 새 구조: address 필드
            if name and 'address' in item:
                coord_map[f'{name}|{item["address"]}'] = (lat, lng)

            # ID 기반 조회도 저장
            if item.get('id'):
                coord_map[item['id']] = (lat, lng)

    elif isinstance(data, dict):
        # 이전 dict 구조가 있을 경우 대비
        pass

    print(f'  기존 좌표 {len(coord_map)}개 로드됨')
    return coord_map


# ──────────────────────────────────────────
# 메인
# ──────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='국토부 실거래가 엑셀 → JSON 변환')
    parser.add_argument('--row-data', default='./rowData', help='엑셀 파일 폴더 (기본값: ./rowData)')
    parser.add_argument('--output', default='./entities/property/model', help='출력 폴더 (기본값: ./entities/property/model)')
    parser.add_argument('--no-geocoding', action='store_true', help='좌표 업데이트(npm run update-coordinates) 건너뜀')
    args = parser.parse_args()

    print('=' * 60)
    print('국토부 실거래가 엑셀 → JSON 변환')
    print('=' * 60)

    # 엑셀 파일 검색 (하위 폴더 포함 재귀 탐색)
    excel_files = glob.glob(os.path.join(args.row_data, '**', '*.xlsx'), recursive=True)
    # 루트 직속 파일도 포함
    excel_files += glob.glob(os.path.join(args.row_data, '*.xlsx'))
    excel_files = sorted(set(excel_files))  # 중복 제거 + 정렬

    if not excel_files:
        print(f'오류: {args.row_data} (하위 폴더 포함)에 xlsx 파일이 없습니다.')
        return 1

    print(f'\n발견된 엑셀 파일 ({len(excel_files)}개):')
    for f in excel_files:
        # rowData 기준 상대 경로 표시
        rel = os.path.relpath(f, args.row_data)
        print(f'  {rel}')

    # 타입별 분류 — 같은 유형 파일이 여러 개(연도별 폴더)면 모두 수집
    typed_files: Dict[str, List[str]] = {}
    for f in excel_files:
        try:
            ftype = detect_file_type(f)
            typed_files.setdefault(ftype, []).append(f)
        except ValueError as e:
            print(f'  경고: {e}')

    print(f'\n타입별 파일 수:')
    for ftype, fpaths in typed_files.items():
        print(f'  {ftype}: {len(fpaths)}개 파일')

    # 출력 디렉터리
    output_dir = Path(args.output)
    props_path = output_dir / 'properties.json'
    prices_path = output_dir / 'prices.json'

    # 기존 좌표 로드
    print(f'\n기존 좌표 데이터 로드 중...')
    existing_coords = load_existing_coords(props_path)

    all_properties: List[dict] = []
    all_prices: List[dict] = []
    all_stats: List[dict] = []

    def read_all(ftype: str, label: str) -> List[dict]:
        """해당 타입의 모든 파일을 읽어 행을 합산 반환"""
        rows: List[dict] = []
        for fpath in typed_files.get(ftype, []):
            rel = os.path.relpath(fpath, args.row_data)
            print(f'  파싱: {rel}')
            rows.extend(read_excel(fpath))
        print(f'  → {label} 합계 {len(rows)}개 행')
        return rows

    # ── 아파트 처리 ──
    if 'apt_sale' in typed_files or 'apt_jeonse' in typed_files:
        if 'apt_sale' not in typed_files:
            print('\n경고: 아파트 매매 파일 없음 → 아파트 데이터 건너뜀')
        else:
            print('\n아파트 매매 데이터 파싱 중...')
            apt_sale_rows = read_all('apt_sale', '아파트 매매')
            print('아파트 전월세 데이터 파싱 중...')
            apt_jeonse_rows = read_all('apt_jeonse', '아파트 전월세')
            print('아파트 데이터 집계 중...')
            props, prices, stats = aggregate(apt_sale_rows, apt_jeonse_rows, 'apartment', existing_coords)
            all_properties.extend(props)
            all_prices.extend(prices)
            stats['type'] = 'apartment'
            all_stats.append(stats)
            print(f'  → {stats["properties"]}개 단지, {stats["units"]}개 유닛')
            print(f'  → 해제 거래 건너뜀: {stats["skipped_cancel"]}건')
            print(f'  → 좌표 없는 신규 단지: {stats["no_coord"]}개')

    # ── 오피스텔 처리 ──
    if 'opt_sale' in typed_files or 'opt_jeonse' in typed_files:
        if 'opt_sale' not in typed_files:
            print('\n경고: 오피스텔 매매 파일 없음 → 오피스텔 데이터 건너뜀')
        else:
            print('\n오피스텔 매매 데이터 파싱 중...')
            opt_sale_rows = read_all('opt_sale', '오피스텔 매매')
            print('오피스텔 전월세 데이터 파싱 중...')
            opt_jeonse_rows = read_all('opt_jeonse', '오피스텔 전월세')
            print('오피스텔 데이터 집계 중...')
            props, prices, stats = aggregate(opt_sale_rows, opt_jeonse_rows, 'officetel', existing_coords)
            all_properties.extend(props)
            all_prices.extend(prices)
            stats['type'] = 'officetel'
            all_stats.append(stats)
            print(f'  → {stats["properties"]}개 단지, {stats["units"]}개 유닛')
            print(f'  → 해제 거래 건너뜀: {stats["skipped_cancel"]}건')
            print(f'  → 좌표 없는 신규 단지: {stats["no_coord"]}개')

    if not all_properties:
        print('\n오류: 처리할 데이터가 없습니다.')
        return 1

    # ── JSON 저장 ──
    print(f'\nJSON 저장 중...')
    output_dir.mkdir(parents=True, exist_ok=True)

    with open(props_path, 'w', encoding='utf-8') as f:
        json.dump(all_properties, f, ensure_ascii=False, indent=2)
    print(f'  → {props_path} 저장 완료')

    with open(prices_path, 'w', encoding='utf-8') as f:
        json.dump(all_prices, f, ensure_ascii=False, indent=2)
    print(f'  → {prices_path} 저장 완료')

    # ── 최종 요약 ──
    total_no_coord = sum(s.get('no_coord', 0) for s in all_stats)
    total_cancel = sum(s.get('skipped_cancel', 0) for s in all_stats)

    print('\n' + '=' * 60)
    print(f'완료! properties: {len(all_properties)}개, prices: {len(all_prices)}개')
    print(f'  해제 거래 건너뜀: {total_cancel}건')
    print(f'  좌표 없는 신규 단지: {total_no_coord}개 (geocoding 필요)')
    print('=' * 60)

    # ── 좌표 업데이트 ──
    if not args.no_geocoding:
        if total_no_coord > 0:
            print('\n좌표 없는 단지가 있습니다. 좌표 업데이트를 시작합니다...')
            print('(KAKAO_REST_API_KEY 환경 변수 필요)')
            os.system('npm run update-coordinates')
        else:
            print('\n모든 단지에 좌표가 있습니다. 좌표 업데이트 건너뜀.')
    else:
        print('\n(--no-geocoding 옵션: 좌표 업데이트 건너뜀)')

    return 0


if __name__ == '__main__':
    exit(main())
